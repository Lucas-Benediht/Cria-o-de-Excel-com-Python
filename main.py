import openpyxl
from colorama import Fore


print(Fore.LIGHTMAGENTA_EX + "Iniciando a criação da planilha")

#criar uma planilha excel
book = openpyxl.Workbook()

#Como visualizar páginas já existentes
print(book.sheetnames)

#Como Criar página
book.create_sheet('Carros')

#Selecionar a página
carros_page = book['Carros']

#Adicionar dados (por linhas!!!)
carros_page.append(['Modelo', 'Ano', 'Valor'])
carros_page.append(['BMWiX', '2023', 'R$:699.950'])
carros_page.append(['BMWiXM60', '2023', 'R$:1.101.950'])
carros_page.append(['BMWi7', '2023', 'R$:1.282.950'])
carros_page.append(['BMWi4', '2023', 'R$:432.950'])
carros_page.append(['BMWiX3', '2023', 'R$:500.950'])

#Salvar a planilha
book.save('Planilha de valores.xlsx')

print(Fore.LIGHTCYAN_EX + 'Criação da planilha finalizada')